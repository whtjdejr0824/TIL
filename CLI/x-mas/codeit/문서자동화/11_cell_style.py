from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
al = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimentions["A"].width = 5

# 1 행의 높이를 50 으로 설정
ws.row_dimentions[1].height = 50

# 스타일 적용
a1.font = Font(color ="FF0000", italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭, 두껍게 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트를 Arial로 설정, 취소선 적용
c1.font = Font(color="0000FF", size=20, underline="single") # 글자 크기를 20, 밑줄 적용

# 테두리 적용
thin_bordar = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_bordar
b1.border = thin_bordar
c1.border = thin_bordar

# 90 점 넘는 셀에 대하여 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대하여 정렬
        cell.alignment = Alignment(horizontal="center",vertical="center")
        # center, left, right, top, bottom

        if cell.column == 1: # A 번호열은 제외
            continue

        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로
            cell.font = Font(color="FF0000") # 폰트 색상 변경


# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

wb.save("sample_style.xlsx")